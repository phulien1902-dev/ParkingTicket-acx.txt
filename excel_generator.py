# excel_generator.py

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelGenerator:

    def __init__(self, output_folder):

        self.output_folder = output_folder

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

    def create(self, tickets):

        wb = Workbook()
        ws = wb.active

        ws.title = "Danh sách vé"

        headers = [
            "STT",
            "Mã vé",
            "Ngày",
            "Giờ",
            "Đơn vị",
            "File QR"
        ]

        # Header
        for col, text in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col)
            cell.value = text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="4F81BD"
            )
            cell.alignment = Alignment(horizontal="center")

        today = datetime.now()

        for index, ticket in enumerate(tickets, start=1):

            ws.cell(index + 1, 1).value = index
            ws.cell(index + 1, 2).value = ticket["code"]
            ws.cell(index + 1, 3).value = today.strftime("%d/%m/%Y")
            ws.cell(index + 1, 4).value = today.strftime("%H:%M:%S")
            ws.cell(index + 1, 5).value = "Cty ABC"
            ws.cell(index + 1, 6).value = ticket["file"]

        # Độ rộng cột
        widths = {
            1: 8,
            2: 18,
            3: 15,
            4: 12,
            5: 35,
            6: 50
        }

        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        excel_file = os.path.join(
            self.output_folder,
            "DanhSachVeGuiXe.xlsx"
        )

        wb.save(excel_file)

        return excel_file